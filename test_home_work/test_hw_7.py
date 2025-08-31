import csv
import io
import os
import zipfile

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
import requests
from selene import browser, query

DIR_PATH = os.path.abspath(os.path.join(os.getcwd(), 'any_files'))


def download_pdf_file():
    browser.open("https://docs.pytest.org/en/stable/tidelift.html")
    download_urls = browser.all("a.reference.external")
    for url in download_urls:
        if url().text == 'PDF Documentation':
            download_url = url.get(query.attribute("href"))
            break
    else:
        raise ValueError("На странице не найдена ссылка на файл")
    content = requests.get(url=download_url, allow_redirects=True).content
    file_name = 'file_hw_7.pdf'
    file_path = os.path.abspath(os.path.join(DIR_PATH, file_name))
    with open(file_path, "wb") as file:
        file.write(content)
    browser.close()


def create_xlsx_file():
    """Документация по openpyxl: https://docs-python.ru/packages/modul-openpyxl/"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPage"
    ws['A1'] = 'Первая ячейка'
    ws['A2'] = 'Вторая ячейка'
    file_name = 'file_hw_7.xlsx'
    file_path = os.path.abspath(os.path.join(DIR_PATH, file_name))
    wb.save(file_path)


def delete_file(file_name: str|list[str, ]):
    if isinstance(file_name, str):
        os.remove(os.path.abspath(os.path.join(DIR_PATH, file_name)))
    else:
        for name in file_name:
            if name.endswith('.csv'):
                continue
            delete_file(name)  # Рекурсивно удаляем файлы из списка кроме csv


def create_archive():
    files_in_dir = [f for f in os.listdir(DIR_PATH) if f.startswith('file_hw_7') and not f.endswith('.zip')]

    file_name = 'file_hw_7.zip'
    file_path = os.path.abspath(os.path.join(DIR_PATH, file_name))
    with zipfile.ZipFile(file_path, 'w') as zip_file:
        for file in files_in_dir:
            add_file = os.path.join(DIR_PATH, file)
            zip_file.write(add_file, os.path.basename(add_file))
    delete_file(files_in_dir)
    return file_name


def test_work_with_zip_files(setup):
    print('')
    download_pdf_file()
    create_xlsx_file()

    zip_file_name = create_archive()
    file_path = os.path.abspath(os.path.join(DIR_PATH, zip_file_name))
    with zipfile.ZipFile(file_path) as zip_file:
        name_files = zip_file.namelist()
        print(name_files)
        for name in name_files:
            file = zip_file.read(name)
            if name.endswith('.csv'):
                reader = list(csv.reader(file.decode('utf-8').splitlines(), delimiter=','))
                second_row = reader[1]
                assert second_row[0] == 'Огурцов'
                assert second_row[1] == 'Петя'
            elif name.endswith('.pdf'):
                reader = PdfReader(io.BytesIO(file))
                pdf_page = reader.pages[0]  # получаем первую страницу
                pdf_text = pdf_page.extract_text()
                assert 'pytest Documentation' in pdf_text
            elif name.endswith('.xlsx'):
                reader = load_workbook(io.BytesIO(file))
                ws = reader.active
                assert ws['A1'].value == 'Первая ячейка'
                assert ws['A2'].value == 'Вторая ячейка'
            else:
                continue
    delete_file(zip_file_name)  # После проверки удаляю архив
